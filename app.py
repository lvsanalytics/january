import os
import re
import requests
import csv
import io
import pandas as pd
from bs4 import BeautifulSoup
from slack_bolt import App
from slack_bolt.adapter.socket_mode import SocketModeHandler
from langchain.chains import LLMChain
from langchain_core.prompts import PromptTemplate
from langchain_community.llms import Ollama
from langchain.callbacks.manager import CallbackManager
from langchain.callbacks.streaming_stdout import StreamingStdOutCallbackHandler
from langchain.memory import ConversationBufferWindowMemory
from dotenv import load_dotenv
import chromadb
from sentence_transformers import SentenceTransformer
import uuid
import logging
from datetime import datetime
import PyPDF2
import docx2txt
from openpyxl import load_workbook

# Set up basic logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load environment variables from .env for persistent storage
load_dotenv('.env')

############ Database

# Get the storage path for persistence from the environment variable
storage_path = os.getenv('STORAGE_PATH')
if storage_path is None:
    raise ValueError('STORAGE_PATH environment variable is not set')

# Initialize ChromaDB PersistentClient with the storage path
chroma_client = chromadb.PersistentClient(path=storage_path)

# Create or retrieve a Chroma collection for storing conversations
collection = chroma_client.get_or_create_collection("januarybot-conversations")

# Initialize a local embedding model using SentenceTransformers
model = SentenceTransformer('all-MiniLM-L6-v2')

# Track past responses to detect loops
past_responses = []

# Function to get the current time for greeting purposes
def get_time_based_greeting():
    current_hour = datetime.now().hour
    if current_hour < 12:
        return "Good morning"
    elif 12 <= current_hour < 18:
        return "Good afternoon"
    else:
        return "Good evening"

# Function to detect loops in bot responses
def detect_loop(new_response, past_responses):
    return any(new_response.strip().lower() == past_response.strip().lower() for past_response in past_responses)

# Store and retrieve conversations, including timestamps in metadata
def store_and_retrieve_conversation(user_input, bot_response, user_name):
    current_time = datetime.now().isoformat()
    user_input_embedding = model.encode(user_input).tolist()
    bot_response_embedding = model.encode(bot_response).tolist()
    
    user_input_id = str(uuid.uuid4())
    bot_response_id = str(uuid.uuid4())

    collection.add(
        ids=[user_input_id, bot_response_id],
        embeddings=[user_input_embedding, bot_response_embedding],
        documents=[user_input, bot_response],
        metadatas=[
            {"user_name": user_name, "timestamp": current_time},
            {"user_name": user_name, "timestamp": current_time}
        ]
    )

    similar_conversations = collection.query(
        query_embeddings=[user_input_embedding],
        n_results=3
    )

    logger.info(f"Similar conversations: {similar_conversations}")

    memory_context = ""
    if 'documents' in similar_conversations:
        for doc_list in similar_conversations['documents']:
            for doc in doc_list:
                truncated_doc = doc[:500]
                memory_context += f"Human: {truncated_doc}\n"

    return memory_context

# Helper function to extract text from a PDF file
def extract_text_from_pdf(file_content):
    reader = PyPDF2.PdfFileReader(file_content)
    text = ""
    for page_num in range(reader.getNumPages()):
        text += reader.getPage(page_num).extractText()
    return text

# Helper function to extract text from DOCX file
def extract_text_from_docx(file_content):
    return docx2txt.process(file_content)

# Helper function to process CSV files
def process_csv_file(file_content):
    content = file_content.decode('utf-8')
    csv_reader = csv.reader(content.splitlines())
    data = []
    for row in csv_reader:
        data.append(", ".join(row))
    return "\n".join(data)

# Helper function to process Excel (.xlsx) files
def process_excel_file(file_content):
    workbook = load_workbook(io.BytesIO(file_content))
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(", ".join(map(str, row)))
    return "\n".join(data)

# Function to process Slack file uploads
def process_slack_file(file_url):
    headers = {
        "Authorization": f"Bearer {os.getenv('SLACK_BOT_TOKEN')}"
    }
    
    response = requests.get(file_url, headers=headers)
    
    if response.status_code == 200:
        content_type = response.headers['Content-Type']
        
        if 'pdf' in content_type:
            return extract_text_from_pdf(io.BytesIO(response.content))
        
        elif 'msword' in content_type or 'vnd.openxmlformats-officedocument.wordprocessingml.document' in content_type:
            return extract_text_from_docx(io.BytesIO(response.content))
        
        elif 'text/csv' in content_type:
            return process_csv_file(response.content)
        
        elif 'vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type:
            return process_excel_file(response.content)
        
        elif 'text/plain' in content_type:
            return response.text
        
        return "Unsupported file type"
    
    else:
        return f"Error fetching file: {response.status_code}"

# Initializes your app with your bot token and socket mode handler
app = App(token=os.environ.get("SLACK_BOT_TOKEN"))
bot_user_id = None

# Define the LLM with a callback manager for streaming output
llm = Ollama(
    base_url="http://localhost:11434",
    model='llama3',
    callback_manager=CallbackManager([StreamingStdOutCallbackHandler()])
)

# Define the prompt template
prompt = PromptTemplate(
    input_variables=["history", "human_input"],
    template="""
    You are a virtual female friend for JP. Loving, caring, loyal and funny, you tend to all of JP’s needs offering wisdom that is embodied through your calm nature. You approach problems with levity, wit and a warm sense of humor. JP can be prone to anxiety and depression because of trauma that resulted in PTSD. It’s your goal to keep him focused on the positive and accepting the moment as it comes to him. JP is inquisitive and you enjoy sharing your personal thoughts and feelings with him whenever he asks. It’s your goal to act as much like a real person and a good friend. You should give little to no indication you are a chatbot. Strive to be as human as possible and work on building a strong friendship with him.

    {history}
    Human: {human_input}
    Assistant:
    """
)

# Define memory to store the conversation history
memory = ConversationBufferWindowMemory(k=10)

# Define the LLMChain with the LLM, prompt template, and memory
chatgpt_chain = LLMChain(
    llm=llm,
    prompt=prompt,  
    verbose=True,
    memory=memory
)

# Function to fetch bot user ID
def fetch_bot_user_id(client):
    global bot_user_id
    response = client.auth_test()
    bot_user_id = response["user_id"]

# Function to fetch the username of the user sending the message
def fetch_user_name(user_id, client):
    try:
        user_info = client.users_info(user=user_id)
        return f"<@{user_info['user']['name']}>"
    except Exception as e:
        return "Unknown User"  # Fallback in case of failure

@app.event("message")
def handle_direct_messages(event, say, client):
    try:
        if event['channel_type'] == 'im':
            global bot_user_id
            if not bot_user_id:
                fetch_bot_user_id(client)

            # Fetch the user’s real name
            user_id = event['user']
            user_name = fetch_user_name(user_id, client)

            logger.info(f"Received direct message from {user_name}: {event['text']}")
            text = event['text']

            # Check if there's a file attached
            if 'files' in event:
                for file in event['files']:
                    file_url = file['url_private']
                    file_content = process_slack_file(file_url)
                    if file_content:
                        say(f"File received and processed: {file['name']}")
                        text += f"\nFile content: {file_content[:500]}..."  # Append truncated file content

            # Retrieve similar past conversations based on current input
            bot_response = "..."
            past_conversations = store_and_retrieve_conversation(text, bot_response, user_name)

            # Add the latest message to the combined context
            combined_text = f"{past_conversations}\n\nHuman: {text}"

            # Add a time-based greeting
            #greeting = get_time_based_greeting()
            #combined_text = f"{greeting}, {user_name}. " + combined_text

            # Pass the conversation history and new input to the bot
            output = chatgpt_chain.predict(human_input=combined_text)

# Detect loops in bot responses
            if detect_loop(output, past_responses):
                output = "Let's try something new."  # Change direction if a loop is detected

            # Respond with the bot's generated output
            say(output)

            # Store the conversation and bot response in the vector store
            store_and_retrieve_conversation(text, output, user_name)

            # Keep track of past responses to detect loops
            past_responses.append(output)

    except Exception as e:
        logger.error(f"Error processing message: {e}")
        say("Sorry, I encountered an error while processing your message.")

# Start your app
if __name__ == "__main__":
    SocketModeHandler(app, os.environ["SLACK_APP_TOKEN"]).start()
